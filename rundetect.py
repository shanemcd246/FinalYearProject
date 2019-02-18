import os
import cv2
import numpy as np
import tensorflow as tf
import sys
import datetime
import uuid
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

#if os.path.exists("/media/student/G00303640/FYP_Excel/WeedDetect.xlsx"):
#  wb = load_workbook("/media/student/G00303640/FYP_Excel/WeedDetect.xlsx")
#  ws = wb.active
#else:
#  wb = Workbook()
#  ws = wb.active
#  ws.append(["Weed", "Accuracy", "Time", "PhotoLoc"])


sys.path.append("..")

from utils import label_map_util
from utils import visualization_utils as vis_util

MODEL_NAME = 'weeds_inference_graph'

CWD_PATH = os.getcwd()

PATH_TO_CKPT = os.path.join(CWD_PATH,'frozen_inference_graph.pb')
PATH_TO_LABELS = os.path.join(CWD_PATH,'object-detection.pbtxt')
NUM_CLASSES = 4

label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
category_index = label_map_util.create_category_index(categories)

detection_graph = tf.Graph()
with detection_graph.as_default():
    od_graph_def = tf.GraphDef()
    with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
        serialized_graph = fid.read()
        od_graph_def.ParseFromString(serialized_graph)
        tf.import_graph_def(od_graph_def, name='')

    sess = tf.Session(graph=detection_graph)

image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')

detection_boxes = detection_graph.get_tensor_by_name('detection_boxes:0')

detection_scores = detection_graph.get_tensor_by_name('detection_scores:0')
detection_classes = detection_graph.get_tensor_by_name('detection_classes:0')

num_detections = detection_graph.get_tensor_by_name('num_detections:0')
video = cv2.VideoCapture('udp://192.168.1.3:12345',cv2.CAP_FFMPEG)
#video.set(cv2.CAP_PROP_FRAME_WIDTH, 800)
#video.set(cv2.CAP_PROP_FRAME_HEIGHT, 600)
while(True):
    ret, frame = video.read()
    frame_expanded = np.expand_dims(frame, axis=0)

    (boxes, scores, classes, num) = sess.run(
        [detection_boxes, detection_scores, detection_classes, num_detections],
        feed_dict={image_tensor: frame_expanded})
    nlist =vis_util.visualize_boxes_and_labels_on_image_array(
        frame,
        np.squeeze(boxes),
        np.squeeze(classes).astype(np.int32),
        np.squeeze(scores),
        category_index,
        use_normalized_coordinates=True,
        line_thickness=2,
        min_score_thresh=0.70)
    nframe = nlist[0]
    name = nlist[1]
    #result = [x.strip() for x in name.split(',')]
    cv2.imshow('Object detector', frame)
    #if "NotSet" not in result:
    #  d = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    #  result.append(str(d))
    #  outfile = '/media/student/G00303640/FYP_Excel/DetectedPhotos/'+result[0]+result[2]+'.jpg'
    #  try:
    #    cv2.imwrite(outfile, nframe)
    #    ws.append([result[0], result[1], result[2], outfile])
    #  except:
    #    print(outfile+ " Failed to save")
    if cv2.waitKey(1) == ord('q'):
        break

#wb.save("/media/student/G00303640/FYP_Excel/WeedDetect.xlsx")
video.release()
cv2.destroyAllWindows()

