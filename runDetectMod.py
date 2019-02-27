import os
import cv2
import numpy as np
import tensorflow as tf
import sys
import datetime
import uuid
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

#if os.path.exists("/media/student/G00303640/FYP_Excel/WeedDetect.xlsx"):
#  wb = load_workbook("/media/student/G00303640/FYP_Excel/WeedDetect.xlsx")
#  ws = wb.active
#else:
#  wb = Workbook()
#  ws = wb.active
#  ws.append(["Weed", "Accuracy", "Time", "PhotoLoc"])


sys.path.append("..")

from utils import label_map_util
from utils import visualization_utils as vis_util

MODEL_NAME = 'weeds_inference_graph'

CWD_PATH = os.getcwd()

PATH_TO_CKPT = os.path.join(CWD_PATH,'frozen_inference_graph.pb')
PATH_TO_LABELS = os.path.join(CWD_PATH,'object-detection.pbtxt')
NUM_CLASSES = 4

label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
category_index = label_map_util.create_category_index(categories)

detection_graph = tf.Graph()
with detection_graph.as_default():
    od_graph_def = tf.GraphDef()
    with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
        serialized_graph = fid.read()
        od_graph_def.ParseFromString(serialized_graph)
        tf.import_graph_def(od_graph_def, name='')

    sess = tf.Session(graph=detection_graph)

image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')

detection_boxes = detection_graph.get_tensor_by_name('detection_boxes:0')

detection_scores = detection_graph.get_tensor_by_name('detection_scores:0')
detection_classes = detection_graph.get_tensor_by_name('detection_classes:0')

num_detections = detection_graph.get_tensor_by_name('num_detections:0')
video = cv2.VideoCapture(1)
#video.set(3, 1920)
#video.set(4, 1080)
while(True):
    ret, frame = video.read()

    blurred_frame = cv2.GaussianBlur(frame,(25,25),0)
    #cv2.imshow("blurred_frame", blurred_frame)

    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

    #Ajust for finer colout detetcion
    sensitivity = 25;
    #Green color
    lower_range = np.array([60 - sensitivity, 60, 30], dtype=np.uint8) 
    upper_range = np.array([60 + sensitivity, 255, 255], dtype=np.uint8)

    #Apply filter
    mask = cv2.inRange(hsv, lower_range, upper_range)

    contours, _ = cv2.findContours(mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE )
    #print(contours);
    x,y,w,h =0,0,0,0
    for c in contours:
        c = max(contours, key = cv2.contourArea)
        x,y,w,h = cv2.boundingRect(c)
        # draw largest conture
        cv2.rectangle(frame,(x,y),(x+w,y+h),(0,255,0),2)
    cv2.drawContours(frame, contours, -1, (255, 255, 0), 1)
    cv2.imshow("frame", frame)
    
    mask = np.zeros(frame.shape,np.uint8)
    mask[y:y+h,x:x+w] = frame[y:y+h,x:x+w]
    cv2.imshow("new back", mask)

    
    frame_expanded = np.expand_dims(mask, axis=0)
    (boxes, scores, classes, num) = sess.run(
        [detection_boxes, detection_scores, detection_classes, num_detections],
        feed_dict={image_tensor: frame_expanded})
    nlist =vis_util.visualize_boxes_and_labels_on_image_array(
        mask,
        np.squeeze(boxes),
        np.squeeze(classes).astype(np.int32),
        np.squeeze(scores),
        category_index,
        use_normalized_coordinates=True,
        line_thickness=2,
        min_score_thresh=0.70)
    nframe = nlist[0]
    name = nlist[1]
    #result = [x.strip() for x in name.split(',')]
    cv2.imshow('Object detector', mask)
    #if "NotSet" not in result:
    #  d = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    #  result.append(str(d))
    #  outfile = '/media/student/G00303640/FYP_Excel/DetectedPhotos/'+result[0]+result[2]+'.jpg'
    #  try:
    #    cv2.imwrite(outfile, nframe)
    #    ws.append([result[0], result[1], result[2], outfile])
    #  except:
    #    print(outfile+ " Failed to save")
    if cv2.waitKey(1) == ord('q'):
        break

#wb.save("/media/student/G00303640/FYP_Excel/WeedDetect.xlsx")
video.release()
cv2.destroyAllWindows()

